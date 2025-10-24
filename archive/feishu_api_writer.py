#!/usr/bin/env python3
"""
Feishu API Writer - Direct Data Transfer
Pushes Excel data directly to Feishu sheet via API
"""

import requests
import json
import openpyxl
import os
from typing import Dict, List, Optional, Any
from datetime import datetime

class FeishuAPIWriter:
    """Handles direct data transfer to Feishu via API"""
    
    def __init__(self, config_file: str = "feishu_config.json"):
        self.config = self.load_config(config_file)
        self.base_url = "https://open.feishu.cn/open-apis"
        self.access_token = None
        
    def load_config(self, config_file: str) -> Dict:
        """Load Feishu configuration"""
        if not os.path.exists(config_file):
            raise FileNotFoundError(f"Config file {config_file} not found")
        
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
                
            required_keys = ['app_id', 'app_secret', 'spreadsheet_token', 'sheet_id']
            for key in required_keys:
                if key not in config:
                    raise ValueError(f"Missing required config key: {key}")
                    
            return config
        except Exception as e:
            raise Exception(f"Error loading config: {e}")
    
    def get_access_token(self) -> Optional[str]:
        """Get Feishu access token"""
        url = f"{self.base_url}/auth/v3/tenant_access_token/internal"
        
        payload = {
            "app_id": self.config['app_id'],
            "app_secret": self.config['app_secret']
        }
        
        try:
            response = requests.post(url, json=payload)
            if response.status_code == 200:
                data = response.json()
                if data.get("code") == 0:
                    self.access_token = data["tenant_access_token"]
                    return self.access_token
                else:
                    print(f"❌ Token error: {data.get('msg')}")
            else:
                print(f"❌ HTTP {response.status_code}: {response.text}")
        except Exception as e:
            print(f"❌ Token request failed: {e}")
        
        return None
    
    def read_excel_data(self, excel_file: str) -> Dict[str, Any]:
        """Read ROAS data from Excel file"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            data = {
                'week': None,
                'overall_roas': None,
                'campaigns': []
            }
            
            # Read headers from row 1
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            print(f"📊 Found headers: {headers}")
            
            # Read data from row 2 (assuming single row format)
            if ws.max_row >= 2:
                # Try to find week number
                week_headers = ['Week', 'week', 'WK', 'Wk']
                for col, header in enumerate(headers):
                    if any(wh in header for wh in week_headers):
                        week_value = ws.cell(row=2, column=col + 1).value
                        if week_value:
                            data['week'] = str(int(week_value) if isinstance(week_value, (int, float)) else week_value)
                            break
                
                # Try to find overall ROAS
                roas_headers = ['Overall ROAS', 'ROAS', 'Overall']
                for col, header in enumerate(headers):
                    if any(rh in header for rh in roas_headers):
                        roas_value = ws.cell(row=2, column=col + 1).value
                        if roas_value:
                            data['overall_roas'] = float(roas_value)
                            break
                
                # Read campaign ROAS values
                for col, header in enumerate(headers):
                    if any(keyword in header.lower() for keyword in ['evergreen', 'prospecting', 'retargeting', 'cns', 's+']):
                        campaign_name = header
                        roas_value = ws.cell(row=2, column=col + 1).value
                        
                        if roas_value is not None and str(roas_value).strip() not in ['', 'N/A', 'n/a']:
                            try:
                                roas_float = float(roas_value)
                                data['campaigns'].append({
                                    'name': campaign_name,
                                    'roas': roas_float
                                })
                            except (ValueError, TypeError):
                                print(f"⚠️  Skipping invalid ROAS value for {campaign_name}: {roas_value}")
            
            print(f"📈 Extracted: Week {data['week']}, Overall ROAS: {data['overall_roas']}")
            print(f"🎯 Found {len(data['campaigns'])} campaigns with ROAS data")
            
            return data
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return {'week': None, 'overall_roas': None, 'campaigns': []}
    
    def get_sheet_metadata(self) -> Optional[Dict]:
        """Get sheet metadata to understand structure"""
        if not self.access_token:
            return None
            
        url = f"{self.base_url}/sheets/v2/spreadsheets/{self.config['spreadsheet_token']}/metainfo"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json"
        }
        
        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                if data.get("code") == 0:
                    return data.get("data", {})
                else:
                    print(f"⚠️  Metadata error: {data.get('msg')}")
            else:
                print(f"❌ HTTP {response.status_code}: {response.text}")
        except Exception as e:
            print(f"❌ Metadata request failed: {e}")
        
        return None
    
    def find_target_range(self, excel_data: Dict) -> Optional[Dict]:
        """Find the target range in Feishu sheet for the data"""
        # This is a simplified version - you'd need to analyze your sheet structure
        # For now, we'll assume a basic structure and return a placeholder
        
        week_num = excel_data.get('week')
        if not week_num:
            print("❌ No week number found in Excel data")
            return None
        
        # You need to implement logic to find:
        # 1. Which row contains the week headers
        # 2. Which column corresponds to the current week
        # 3. Which rows contain the campaign names
        
        # Placeholder return - you'll need to customize this
        return {
            'week_row': 1,  # Row containing week numbers
            'campaign_start_row': 2,  # First row with campaign data
            'week_column': int(week_num) + 2,  # Simplified - you'll need to map this
            'campaign_names': [c['name'] for c in excel_data['campaigns']]
        }
    
    def write_data_to_sheet(self, excel_data: Dict) -> bool:
        """Write data to Feishu sheet"""
        if not self.access_token:
            print("❌ No access token available")
            return False
        
        target_range = self.find_target_range(excel_data)
        if not target_range:
            print("❌ Could not determine target range")
            return False
        
        # Build the data to write
        values = []
        
        # Add overall ROAS if available
        if excel_data['overall_roas']:
            values.append([excel_data['overall_roas']])
        
        # Add campaign ROAS values
        for campaign in excel_data['campaigns']:
            values.append([campaign['roas']])
        
        # For now, let's do a simple test write to cell A1
        # You'll need to customize the range based on your sheet structure
        range_str = f"{self.config['sheet_id']}!A1:A{len(values)}"
        
        url = f"{self.base_url}/sheets/v2/spreadsheets/{self.config['spreadsheet_token']}/values"
        
        payload = {
            "valueRange": {
                "range": range_str,
                "values": values
            }
        }
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json"
        }
        
        try:
            print(f"📝 Writing data to range: {range_str}")
            response = requests.put(url, json=payload, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                if data.get("code") == 0:
                    print("✅ Data successfully written to Feishu sheet!")
                    return True
                else:
                    print(f"❌ Write error: {data.get('msg')}")
            else:
                print(f"❌ HTTP {response.status_code}: {response.text}")
                
        except Exception as e:
            print(f"❌ Write request failed: {e}")
        
        return False
    
    def process_and_upload(self, excel_file: str) -> bool:
        """Complete workflow: read Excel and upload to Feishu"""
        print(f"🚀 Starting Feishu upload process...")
        
        # Step 1: Get access token
        print("🔑 Getting access token...")
        if not self.get_access_token():
            return False
        
        # Step 2: Read Excel data
        print(f"📖 Reading Excel file: {excel_file}")
        excel_data = self.read_excel_data(excel_file)
        
        if not excel_data['campaigns']:
            print("❌ No campaign data found in Excel file")
            return False
        
        # Step 3: Write to Feishu
        print("📝 Writing to Feishu sheet...")
        return self.write_data_to_sheet(excel_data)

def main():
    """Main function to test the Feishu API writer"""
    
    # Check for Excel file
    excel_files = [
        "week_37_feishu_ready.xlsx",
        "week_37_AB_format.xlsx",
        "week_37_roas_data.xlsx"
    ]
    
    excel_file = None
    for file in excel_files:
        if os.path.exists(file):
            excel_file = file
            break
    
    if not excel_file:
        print("❌ No Excel output file found. Please run the main generator first.")
        print("💡 Try: python ab_format_excel_generator.py")
        return
    
    print(f"📊 Using Excel file: {excel_file}")
    
    # Create writer and process
    try:
        writer = FeishuAPIWriter()
        success = writer.process_and_upload(excel_file)
        
        if success:
            print("\n🎉 Successfully uploaded data to Feishu!")
        else:
            print("\n❌ Upload failed. Check the error messages above.")
            
    except Exception as e:
        print(f"\n❌ Error: {e}")

if __name__ == "__main__":
    main()