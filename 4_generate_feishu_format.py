#!/usr/bin/env python3
"""
Excel Generator for Lululemon Adobe ROAS Data - CLEANED VERSION
Simplified version focused on core Excel generation without Feishu API complexity
Creates Excel files for manual copy-paste workflow
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
from typing import Dict, List, Optional
from parsers import CampaignMapper

class FeishuExcelGenerator:
    """Generates Excel files optimized for manual copy-paste workflow"""
    
    def __init__(self):
        self.campaign_mapper = CampaignMapper()
        print("🧹 Cleaned FeishuExcelGenerator initialized - Excel-only mode")
        
    def extract_overall_roas(self, email_text: str) -> Optional[float]:
        """Extract overall ROAS from email summary"""
        # Look for pattern like "Week 37 drove a 3.24x ROAS"
        roas_match = re.search(r'Week\s+\d+.*?drove\s+a\s+(\d+\.?\d*)x\s+ROAS', email_text, re.IGNORECASE)
        if roas_match:
            return float(roas_match.group(1))
        
        # Fallback: look for any ROAS pattern
        roas_match = re.search(r'(\d+\.?\d*)x\s+ROAS', email_text, re.IGNORECASE)
        return float(roas_match.group(1)) if roas_match else None
    
    def extract_week_number(self, email_text: str) -> str:
        """Extract week number"""
        week_match = re.search(r'WK\s+(\d+)', email_text, re.IGNORECASE)
        return week_match.group(1) if week_match else "Unknown"
    
    def parse_campaigns(self, email_text: str) -> List[Dict]:
        """Parse individual campaigns"""
        campaigns = []
        lines = email_text.strip().split('\n')
        
        in_campaign_section = False
        current_campaign = {}
        
        for line in lines:
            line = line.strip()
            
            if "Campaign Name" in line:
                in_campaign_section = True
                continue
                
            if not in_campaign_section:
                continue
                
            if "TOTAL" in line:
                break
                
            if "EVRGN" in line and not current_campaign:
                current_campaign['raw_campaign_name'] = line
                continue
                
            if current_campaign and 'spend' not in current_campaign:
                if line.startswith('$') and ',' in line:
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
                elif line.startswith('$') and line.replace('$', '').replace(',', '').replace('.', '').isdigit():
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
            
            if current_campaign and 'roas' not in current_campaign:
                if re.match(r'^\d+\.\d+$', line):
                    current_campaign['roas'] = float(line)
                    
                    if self.validate_campaign_data(current_campaign):
                        processed_campaign = self.process_campaign_data(current_campaign)
                        if processed_campaign:
                            campaigns.append(processed_campaign)
                    
                    current_campaign = {}
                    continue
        
        return campaigns
    
    def validate_campaign_data(self, campaign: Dict) -> bool:
        """Validate campaign data"""
        required_fields = ['raw_campaign_name', 'spend', 'roas']
        return all(field in campaign for field in required_fields)
    
    def process_campaign_data(self, campaign: Dict) -> Optional[Dict]:
        """Process campaign data with mapping"""
        try:
            mapped_campaign_name = self.campaign_mapper.map_email_campaign_to_csv(
                campaign['raw_campaign_name']
            )
            
            if not mapped_campaign_name:
                return None
            
            return {
                'target_campaign_name': mapped_campaign_name,
                'raw_campaign_name': campaign['raw_campaign_name'],
                'spend': campaign['spend'],
                'roas': campaign['roas']
            }
            
        except (ValueError, AttributeError):
            return None
    
    def get_target_campaign_order(self) -> List[str]:
        """Return the exact order of campaigns as they appear in the target tracker"""
        return [
            "US Evergreen Prospecting (Lowest Cost)",
            "S+ 2.0 US Evergreen Prospecting (Lowest Cost)",
            "US Evergreen Prospecting (Cost Cap)",
            "US Evergreen Retargeting (Lowest Cost)",
            "Canada Evergreen Prospecting (Lowest Cost)",
            "S+ 2.0 Canada Evergreen Prospecting (Lowest Cost)",
            "Canada Evergreen Retargeting (Lowest Cost)",
            "US CNS Manual (carousel ads only - Lowest Cost)",
            "US CNS S+ 2.0 (carousel ads only - Lowest Cost)",
            "Canada CNS Manual (carousel ads only - Lowest Cost)",
            "Canada CNS S+ 2.0 (carousel ads only - Lowest Cost)"
        ]
    
    def create_feishu_excel(self, email_text: str, output_filename: str = None) -> str:
        """Create Excel file optimized for manual copy-paste workflow"""
        
        if not output_filename:
            week = self.extract_week_number(email_text)
            output_filename = f"week_{week}_feishu_ready.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feishu Ready Data"
        
        # Get data
        week = self.extract_week_number(email_text)
        overall_roas = self.extract_overall_roas(email_text)
        campaigns = self.parse_campaigns(email_text)
        
        # Create campaign lookup
        campaign_lookup = {campaign['target_campaign_name']: campaign for campaign in campaigns}
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center")
        
        currency_format = '"$"#,##0'
        roas_format = '0.00"x"'
        
        # Headers - matching target tracker format
        headers = ["Week", "Overall ROAS"] + self.get_target_campaign_order()
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data row
        row = 2
        
        # Week number
        ws.cell(row=row, column=1, value=int(week))
        ws.cell(row=row, column=1).font = data_font
        ws.cell(row=row, column=1).alignment = data_alignment
        
        # Overall ROAS
        if overall_roas:
            ws.cell(row=row, column=2, value=overall_roas)
            ws.cell(row=row, column=2).font = data_font
            ws.cell(row=row, column=2).alignment = data_alignment
            ws.cell(row=row, column=2).number_format = roas_format
        
        # Campaign ROAS values - in exact target order
        for col, campaign_name in enumerate(self.get_target_campaign_order(), 3):
            if campaign_name in campaign_lookup:
                campaign = campaign_lookup[campaign_name]
                ws.cell(row=row, column=col, value=campaign['roas'])
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = data_alignment
                ws.cell(row=row, column=col).number_format = roas_format
            else:
                # Campaign not found - leave blank (N/A)
                ws.cell(row=row, column=col, value="N/A")
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = data_alignment
        
        # Set column widths for optimal copy-paste
        ws.column_dimensions['A'].width = 8   # Week
        ws.column_dimensions['B'].width = 12  # Overall ROAS
        
        # Campaign columns - standard width for easy copying
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Cleaned version: Removed complex borders for simplicity
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save file
        wb.save(output_filename)
        return output_filename
    
    def create_detailed_sheet(self, email_text: str, output_filename: str = None) -> str:
        """Create additional sheet with detailed campaign info for reference"""
        
        if not output_filename:
            week = self.extract_week_number(email_text)
            output_filename = f"week_{week}_detailed.xlsx"
        
        # Create main Feishu-ready file first
        main_filename = self.create_feishu_excel(email_text, output_filename)
        
        # Reopen and add detailed sheet
        wb = openpyxl.load_workbook(main_filename)
        
        # Create detailed sheet
        ws_detail = wb.create_sheet("Detailed Campaign Data")
        
        # Headers for detailed sheet
        headers = ['Week', 'Campaign Name', 'Spend', 'ROAS', 'Raw Campaign Name']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get detailed campaign data
        week = self.extract_week_number(email_text)
        campaigns = self.parse_campaigns(email_text)
        
        # Write detailed data
        for row, campaign in enumerate(campaigns, 2):
            ws_detail.cell(row=row, column=1, value=int(week))
            ws_detail.cell(row=row, column=2, value=campaign['target_campaign_name'])
            ws_detail.cell(row=row, column=3, value=campaign['spend'])
            ws_detail.cell(row=row, column=4, value=campaign['roas'])
            ws_detail.cell(row=row, column=5, value=campaign['raw_campaign_name'])
            
            # Formatting
            ws_detail.cell(row=row, column=3).number_format = '"$"#,##0'
            ws_detail.cell(row=row, column=4).number_format = '0.00"x"'
        
        # Set column widths
        ws_detail.column_dimensions['A'].width = 8
        ws_detail.column_dimensions['B'].width = 45
        ws_detail.column_dimensions['C'].width = 15
        ws_detail.column_dimensions['D'].width = 10
        ws_detail.column_dimensions['E'].width = 40
        
        # Save with both sheets
        wb.save(output_filename)
        return output_filename

def main():
    """Generate cleaned Excel files for manual workflow"""
    
    # Read the WK 37 email file
    with open('input/week_37_email_input.txt', 'r') as f:
        email_text = f.read()
    
    # Generate Excel file
    generator = FeishuExcelGenerator()
    
    # Create main file (single row for copy-paste)
    main_filename = generator.create_feishu_excel(email_text)
    
    print(f"✅ Excel file created for manual copy-paste:")
    print(f"📊 File: {main_filename}")
    print(f"📈 Overall ROAS: {generator.extract_overall_roas(email_text)}x")
    print(f"\n🎯 Ready for manual copy-paste into Feishu tracker")

if __name__ == "__main__":
    main()