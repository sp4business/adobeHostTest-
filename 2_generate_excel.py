#!/usr/bin/env python3
"""
Excel Generator for Lululemon Adobe ROAS Data
Creates properly formatted Excel files for easy Feishu sheet import
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
from typing import Dict, List, Optional
from parsers import CampaignMapper

class ExcelGenerator:
    """Generates Excel files with proper formatting for Feishu sheets"""
    
    def __init__(self):
        self.campaign_mapper = CampaignMapper()
        
    def extract_overall_totals(self, email_text: str) -> Dict:
        """Extract overall totals from email"""
        lines = email_text.strip().split('\n')
        
        totals = {
            'total_spend': None,
            'total_revenue': None,
            'total_roas': None,
            'adjusted_spend': None,
            'adjusted_revenue': None,
            'adjusted_roas': None
        }
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            if line == "TOTAL":
                if i + 3 < len(lines):
                    try:
                        spend_line = lines[i + 1].strip()
                        revenue_line = lines[i + 2].strip()
                        roas_line = lines[i + 3].strip()
                        
                        if spend_line.startswith('$'):
                            totals['total_spend'] = float(spend_line.replace('$', '').replace(',', ''))
                        if revenue_line.startswith('$'):
                            totals['total_revenue'] = float(revenue_line.replace('$', '').replace(',', ''))
                        if re.match(r'^\d+\.?\d*$', roas_line):
                            totals['total_roas'] = float(roas_line)
                    except (ValueError, IndexError):
                        pass
            
            if "Total (minus credit + lag revenue)" in line:
                if i + 3 < len(lines):
                    try:
                        adj_spend_line = lines[i + 1].strip()
                        adj_revenue_line = lines[i + 2].strip()
                        adj_roas_line = lines[i + 3].strip()
                        
                        if adj_spend_line.startswith('$'):
                            totals['adjusted_spend'] = float(adj_spend_line.replace('$', '').replace(',', ''))
                        if adj_revenue_line.startswith('$'):
                            totals['adjusted_revenue'] = float(adj_revenue_line.replace('$', '').replace(',', ''))
                        if re.match(r'^\d+\.?\d*$', adj_roas_line):
                            totals['adjusted_roas'] = float(adj_roas_line)
                    except (ValueError, IndexError):
                        pass
        
        return totals
    
    def extract_week_number(self, email_text: str) -> str:
        """Extract week number"""
        week_match = re.search(r'WK\s+(\d+)', email_text, re.IGNORECASE)
        return week_match.group(1) if week_match else "Unknown"
    
    def parse_campaigns(self, email_text: str) -> List[Dict]:
        """Parse individual campaigns"""
        campaigns = []
        lines = email_text.strip().split('\n')
        
        in_campaign_section = False
        current_campaign = {}
        
        for line in lines:
            line = line.strip()
            
            if "Campaign Name" in line:
                in_campaign_section = True
                continue
                
            if not in_campaign_section:
                continue
                
            if "TOTAL" in line:
                break
                
            if "EVRGN" in line and not current_campaign:
                current_campaign['raw_campaign_name'] = line
                continue
                
            if current_campaign and 'spend' not in current_campaign:
                if line.startswith('$') and ',' in line:
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
                elif line.startswith('$') and line.replace('$', '').replace(',', '').replace('.', '').isdigit():
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
            
            if current_campaign and 'revenue' not in current_campaign:
                if line.startswith('$') and ',' in line and 'spend' in current_campaign:
                    current_campaign['revenue'] = float(line.replace('$', '').replace(',', ''))
                    continue
                elif line.startswith('$') and line.replace('$', '').replace(',', '').replace('.', '').isdigit() and 'spend' in current_campaign:
                    current_campaign['revenue'] = float(line.replace('$', '').replace(',', ''))
                    continue
            
            if current_campaign and 'roas' not in current_campaign:
                if re.match(r'^\d+\.\d+$', line):
                    current_campaign['roas'] = float(line)
                    
                    if self.validate_campaign_data(current_campaign):
                        processed_campaign = self.process_campaign_data(current_campaign)
                        if processed_campaign:
                            campaigns.append(processed_campaign)
                    
                    current_campaign = {}
                    continue
        
        return campaigns
    
    def validate_campaign_data(self, campaign: Dict) -> bool:
        """Validate campaign data"""
        required_fields = ['raw_campaign_name', 'spend', 'revenue', 'roas']
        return all(field in campaign for field in required_fields)
    
    def process_campaign_data(self, campaign: Dict) -> Optional[Dict]:
        """Process campaign data with mapping"""
        try:
            mapped_campaign_name = self.campaign_mapper.map_email_campaign_to_csv(
                campaign['raw_campaign_name']
            )
            
            if not mapped_campaign_name:
                return None
            
            return {
                'target_campaign_name': mapped_campaign_name,
                'raw_campaign_name': campaign['raw_campaign_name'],
                'spend': campaign['spend'],
                'revenue': campaign['revenue'],
                'roas': campaign['roas']
            }
            
        except (ValueError, AttributeError):
            return None
    
    def create_excel_file(self, email_text: str, output_filename: str = None) -> str:
        """Create Excel file with formatted data"""
        
        if not output_filename:
            week = self.extract_week_number(email_text)
            output_filename = f"week_{week}_roas_data.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main data sheet
        ws_data = wb.create_sheet("Campaign Data")
        self.create_data_sheet(ws_data, email_text)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        self.create_summary_sheet(ws_summary, email_text)
        
        # Save file
        wb.save(output_filename)
        return output_filename
    
    def create_data_sheet(self, worksheet, email_text: str):
        """Create main campaign data sheet"""
        
        week = self.extract_week_number(email_text)
        campaigns = self.parse_campaigns(email_text)
        
        # Headers
        headers = ['Week', 'Campaign Name', 'Spend', 'Revenue', 'ROAS', 'Raw Campaign Name']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, campaign in enumerate(campaigns, 2):
            worksheet.cell(row=row, column=1, value=int(week))
            worksheet.cell(row=row, column=2, value=campaign['target_campaign_name'])
            worksheet.cell(row=row, column=3, value=campaign['spend'])
            worksheet.cell(row=row, column=4, value=campaign['revenue'])
            worksheet.cell(row=row, column=5, value=campaign['roas'])
            worksheet.cell(row=row, column=6, value=campaign['raw_campaign_name'])
        
        # Format columns
        worksheet.column_dimensions['A'].width = 8   # Week
        worksheet.column_dimensions['B'].width = 50  # Campaign Name
        worksheet.column_dimensions['C'].width = 15  # Spend
        worksheet.column_dimensions['D'].width = 15  # Revenue
        worksheet.column_dimensions['E'].width = 10  # ROAS
        worksheet.column_dimensions['F'].width = 40  # Raw Campaign Name
        
        # Number formatting
        for row in range(2, len(campaigns) + 2):
            worksheet.cell(row=row, column=3).number_format = '"$"#,##0'  # Spend
            worksheet.cell(row=row, column=4).number_format = '"$"#,##0'  # Revenue
            worksheet.cell(row=row, column=5).number_format = '0.00"x"'    # ROAS
    
    def create_summary_sheet(self, worksheet, email_text: str):
        """Create summary sheet with overall totals"""
        
        week = self.extract_week_number(email_text)
        totals = self.extract_overall_totals(email_text)
        
        # Title
        title_font = Font(size=16, bold=True)
        worksheet.cell(row=1, column=1, value=f"Week {week} ROAS Summary")
        worksheet.cell(row=1, column=1).font = title_font
        
        # Headers
        header_font = Font(bold=True)
        
        # Overall Totals section
        worksheet.cell(row=3, column=1, value="Overall Totals").font = header_font
        worksheet.cell(row=4, column=1, value="Metric")
        worksheet.cell(row=4, column=2, value="Value")
        
        row = 5
        if totals['total_spend']:
            worksheet.cell(row=row, column=1, value="Total Spend")
            worksheet.cell(row=row, column=2, value=totals['total_spend'])
            worksheet.cell(row=row, column=2).number_format = '"$"#,##0'
            row += 1
        
        if totals['total_revenue']:
            worksheet.cell(row=row, column=1, value="Total Revenue")
            worksheet.cell(row=row, column=2, value=totals['total_revenue'])
            worksheet.cell(row=row, column=2).number_format = '"$"#,##0'
            row += 1
        
        if totals['total_roas']:
            worksheet.cell(row=row, column=1, value="Total ROAS")
            worksheet.cell(row=row, column=2, value=totals['total_roas'])
            worksheet.cell(row=row, column=2).number_format = '0.00"x"'
            row += 1
        
        # Adjusted Totals section
        row += 1
        worksheet.cell(row=row, column=1, value="Adjusted Totals (minus credit + lag)").font = header_font
        row += 1
        
        if totals['adjusted_spend']:
            worksheet.cell(row=row, column=1, value="Adjusted Spend")
            worksheet.cell(row=row, column=2, value=totals['adjusted_spend'])
            worksheet.cell(row=row, column=2).number_format = '"$"#,##0'
            row += 1
        
        if totals['adjusted_revenue']:
            worksheet.cell(row=row, column=1, value="Adjusted Revenue")
            worksheet.cell(row=row, column=2, value=totals['adjusted_revenue'])
            worksheet.cell(row=row, column=2).number_format = '"$"#,##0'
            row += 1
        
        if totals['adjusted_roas']:
            worksheet.cell(row=row, column=1, value="Adjusted ROAS")
            worksheet.cell(row=row, column=2, value=totals['adjusted_roas'])
            worksheet.cell(row=row, column=2).number_format = '0.00"x"'
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """Generate Excel file for Week 37"""
    
    # Read the WK 37 email file
    with open('input/week_37_email_input.txt', 'r') as f:
        email_text = f.read()
    
    # Generate Excel file
    generator = ExcelGenerator()
    filename = generator.create_excel_file(email_text)
    
    print(f"✅ Excel file created: {filename}")
    print("📊 Ready for Feishu sheet import!")
    print("\nFeatures:")
    print("- Properly formatted campaign names")
    print("- Smart+ 2.0 and CNS mapping applied")
    print("- Currency formatting ($)")
    print("- ROAS formatting (x)")
    print("- Summary sheet with totals")
    print("- Two sheets: 'Campaign Data' and 'Summary'")

if __name__ == "__main__":
    main()