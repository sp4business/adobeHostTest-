#!/usr/bin/env python3
"""
Excel Generator for Lululemon Adobe ROAS Data - A/B Format
Creates Excel files with A/B column format for easy copy-paste into Feishu
Column A: Campaign Name, Column B: ROAS Value
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
from typing import Dict, List, Optional
from parsers import CampaignMapper

class ABFormatExcelGenerator:
    """Generates Excel files in A/B format for Feishu copy-paste"""
    
    def __init__(self):
        self.campaign_mapper = CampaignMapper()
        
    def extract_overall_roas(self, email_text: str) -> Optional[float]:
        """Extract overall ROAS from email summary"""
        # Look for pattern like "Week 37 drove a 3.24x ROAS"
        roas_match = re.search(r'Week\s+\d+.*?drove\s+a\s+(\d+\.?\d*)x\s+ROAS', email_text, re.IGNORECASE)
        if roas_match:
            return float(roas_match.group(1))
        
        # Fallback: look for any ROAS pattern
        roas_match = re.search(r'(\d+\.?\d*)x\s+ROAS', email_text, re.IGNORECASE)
        return float(roas_match.group(1)) if roas_match else None
    
    def extract_week_number(self, email_text: str) -> str:
        """Extract week number"""
        week_match = re.search(r'WK\s+(\d+)', email_text, re.IGNORECASE)
        return week_match.group(1) if week_match else "Unknown"
    
    def parse_campaigns(self, email_text: str) -> List[Dict]:
        """Parse individual campaigns"""
        campaigns = []
        lines = email_text.strip().split('\n')
        
        in_campaign_section = False
        current_campaign = {}
        
        for line in lines:
            line = line.strip()
            
            if "Campaign Name" in line:
                in_campaign_section = True
                continue
                
            if not in_campaign_section:
                continue
                
            if "TOTAL" in line:
                break
                
            if "EVRGN" in line and not current_campaign:
                current_campaign['raw_campaign_name'] = line
                continue
                
            if current_campaign and 'spend' not in current_campaign:
                if line.startswith('$') and ',' in line:
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
                elif line.startswith('$') and line.replace('$', '').replace(',', '').replace('.', '').isdigit():
                    current_campaign['spend'] = float(line.replace('$', '').replace(',', ''))
                    continue
            
            if current_campaign and 'roas' not in current_campaign:
                if re.match(r'^\d+\.\d+$', line):
                    current_campaign['roas'] = float(line)
                    
                    if self.validate_campaign_data(current_campaign):
                        processed_campaign = self.process_campaign_data(current_campaign)
                        if processed_campaign:
                            campaigns.append(processed_campaign)
                    
                    current_campaign = {}
                    continue
        
        return campaigns
    
    def validate_campaign_data(self, campaign: Dict) -> bool:
        """Validate campaign data"""
        required_fields = ['raw_campaign_name', 'spend', 'roas']
        return all(field in campaign for field in required_fields)
    
    def process_campaign_data(self, campaign: Dict) -> Optional[Dict]:
        """Process campaign data with mapping"""
        try:
            mapped_campaign_name = self.campaign_mapper.map_email_campaign_to_csv(
                campaign['raw_campaign_name']
            )
            
            if not mapped_campaign_name:
                return None
            
            return {
                'target_campaign_name': mapped_campaign_name,
                'raw_campaign_name': campaign['raw_campaign_name'],
                'spend': campaign['spend'],
                'roas': campaign['roas']
            }
            
        except (ValueError, AttributeError):
            return None
    
    def get_target_campaign_order(self) -> List[str]:
        """Return the exact order of campaigns as they appear in the target tracker"""
        return [
            "US Evergreen Prospecting (Lowest Cost)",
            "S+ 2.0 US Evergreen Prospecting (Lowest Cost)",
            "US Evergreen Prospecting (Cost Cap)",
            "US Evergreen Retargeting (Lowest Cost)",
            "Canada Evergreen Prospecting (Lowest Cost)",
            "S+ 2.0 Canada Evergreen Prospecting (Lowest Cost)",
            "Canada Evergreen Retargeting (Lowest Cost)",
            "US CNS Manual (carousel ads only - Lowest Cost)",
            "US CNS S+ 2.0 (carousel ads only - Lowest Cost)",
            "Canada CNS Manual (carousel ads only - Lowest Cost)",
            "Canada CNS S+ 2.0 (carousel ads only - Lowest Cost)"
        ]
    
    def create_ab_format_excel(self, email_text: str, output_filename: str = None) -> str:
        """Create Excel file in A/B format for Feishu copy-paste"""
        
        if not output_filename:
            week = self.extract_week_number(email_text)
            output_filename = f"week_{week}_AB_format.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AB Format - Feishu Ready"
        
        # Get data
        week = self.extract_week_number(email_text)
        overall_roas = self.extract_overall_roas(email_text)
        campaigns = self.parse_campaigns(email_text)
        
        # Create campaign lookup
        campaign_lookup = {campaign['target_campaign_name']: campaign for campaign in campaigns}
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center")
        
        roas_format = '0.00"x"'
        
        # Headers - A/B format
        ws.cell(row=1, column=1, value="Campaign Name")
        ws.cell(row=1, column=2, value="ROAS")
        
        # Style headers
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data in A/B format - Overall ROAS first
        row = 2
        
        # Overall ROAS (Week X)
        if overall_roas:
            ws.cell(row=row, column=1, value=f"Week {week} Overall ROAS")
            ws.cell(row=row, column=2, value=overall_roas)
            ws.cell(row=row, column=2).number_format = roas_format
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = data_font
            ws.cell(row=row, column=2).alignment = data_alignment
            row += 1
        
        # Add separator row
        row += 1
        
        # Campaign ROAS values - in exact target order
        for campaign_name in self.get_target_campaign_order():
            if campaign_name in campaign_lookup:
                campaign = campaign_lookup[campaign_name]
                ws.cell(row=row, column=1, value=campaign_name)
                ws.cell(row=row, column=2, value=campaign['roas'])
                ws.cell(row=row, column=2).number_format = roas_format
            else:
                # Campaign not found - leave ROAS blank
                ws.cell(row=row, column=1, value=campaign_name)
                ws.cell(row=row, column=2, value="")  # Blank for missing campaigns
            
            # Style data cells
            ws.cell(row=row, column=1).font = data_font
            ws.cell(row=row, column=2).font = data_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2).alignment = data_alignment
            
            row += 1
        
        # Set column widths for optimal copy-paste
        ws.column_dimensions['A'].width = 50  # Campaign names need more space
        ws.column_dimensions['B'].width = 12  # ROAS values
        
        # Optional: Add borders for better visibility
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        for row_num in range(1, row):
            for col_num in range(1, 3):
                ws.cell(row=row_num, column=col_num).border = thin_border
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save file
        wb.save(output_filename)
        return output_filename
    
    def create_summary_sheet(self, email_text: str, output_filename: str = None) -> str:
        """Create additional summary sheet for reference"""
        
        if not output_filename:
            week = self.extract_week_number(email_text)
            output_filename = f"week_{week}_AB_summary.xlsx"
        
        # Create main A/B file first
        main_filename = self.create_ab_format_excel(email_text, output_filename)
        
        # Reopen and add summary sheet
        wb = openpyxl.load_workbook(main_filename)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary & Notes")
        
        # Get data
        week = self.extract_week_number(email_text)
        overall_roas = self.extract_overall_roas(email_text)
        campaigns = self.parse_campaigns(email_text)
        
        # Summary data
        ws_summary.cell(row=1, column=1, value=f"Week {week} Summary")
        ws_summary.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        ws_summary.cell(row=3, column=1, value="Overall ROAS:")
        ws_summary.cell(row=3, column=2, value=f"{overall_roas}x" if overall_roas else "N/A")
        ws_summary.cell(row=3, column=2).font = Font(bold=True)
        
        ws_summary.cell(row=4, column=1, value="Total Campaigns:")
        ws_summary.cell(row=4, column=2, value=len(campaigns))
        
        ws_summary.cell(row=5, column=1, value="Week Number:")
        ws_summary.cell(row=5, column=2, value=week)
        
        # Notes section
        ws_summary.cell(row=7, column=1, value="Notes:")
        ws_summary.cell(row=7, column=1).font = Font(bold=True)
        ws_summary.cell(row=8, column=1, value="- Copy Column B from 'A/B Format - Feishu Ready' sheet")
        ws_summary.cell(row=9, column=1, value="- Paste into Feishu tracker starting at appropriate week column")
        ws_summary.cell(row=10, column=1, value="- N/A values indicate campaigns not active this week")
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 15
        
        # Save with both sheets
        wb.save(output_filename)
        return output_filename

def main():
    """Generate A/B format Excel file"""
    
    # Read the WK 37 email file
    with open('input/week_37_email_input.txt', 'r') as f:
        email_text = f.read()
    
    # Generate A/B format Excel file
    generator = ABFormatExcelGenerator()
    
    # Create main A/B format file
    main_filename = generator.create_ab_format_excel(email_text)
    
    # Create summary file (with both sheets)
    summary_filename = generator.create_summary_sheet(email_text)
    
    print(f"✅ A/B Format Excel files created:")
    print(f"📊 Main A/B format (copy-paste ready): {main_filename}")
    print(f"📋 Summary file (with instructions): {summary_filename}")
    print(f"\n🎯 A/B Format: Column A = Campaign Name, Column B = ROAS")
    print(f"📈 Overall ROAS: {generator.extract_overall_roas(email_text)}x")
    print(f"\n💡 Usage: Copy Column B from 'A/B Format - Feishu Ready' sheet")
    print(f"   and paste directly into your Feishu tracker!")

if __name__ == "__main__":
    main()